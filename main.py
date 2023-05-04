import openpyxl
from scipy.signal import find_peaks
from openpyxl import Workbook
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt

import Battery


def getDictionaryFromxlsl(file_location, file_name, sheet_name):
    xlsx_file = Path(file_location, file_name)

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet = wb_obj[sheet_name]

    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)

    data = {}

    for name in col_names:
        data[name] = []

    first = True
    for row in sheet.iter_rows(max_row=sheet.max_row):
        i = 0
        for cell in row:
            if not first:
                data[col_names[i]].append(cell.value)
            i += 1
        if first:
            first = False
    return data


def returnPeaksAndValleysAndShowChart(data, chart_name):
    names = list(data.keys())
    x = np.array(data[names[0]])
    y = np.array(data[names[1]])

    mean = np.mean(y)
    std = np.std(y)

    peaks, _ = find_peaks(y, height=mean)
    valleys, _ = find_peaks(-y, height=-(mean - std))

    showChart(x, y, peaks, valleys, chart_name)

    return peaks, valleys


def showChart(x, y, peaks, valleys, chart_name):
    plt.plot(x, y)
    plt.plot(x[peaks], y[peaks], "p", label="szczyty")
    plt.plot(x[valleys], y[valleys], "v", label="doliny")

    plt.savefig(f"Charts/{chart_name}.png", dpi=300)
    plt.legend()
    plt.show()


def main():
    sheet_name = "18.01.2022"
    data = getDictionaryFromxlsl("Dane", "daneX.xlsx", sheet_name)
    # data = getDictionaryFromxlsl("Dane", "daneX.xlsx", "17-21.01.2022")
    # noc od 22 do 8 rano, od 9 do 21 to dzień
    godziny = 24
    pojemnosc_magazynu = 2000

    # magazyn = Battery.Battery(pojemnosc_magazynu)
    # magazyn.change_current_charge(1000)

    peaks, valleys = returnPeaksAndValleysAndShowChart(data, sheet_name)


main()
